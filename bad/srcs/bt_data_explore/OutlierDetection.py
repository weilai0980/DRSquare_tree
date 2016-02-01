import openpyxl as px
import random as rnd
import numpy as np
import heapq
import pylab as plt
#import matplotlib as mplt
import math
import sys

def isNumber(s):
    try:
        float(s)
        return True
    except ValueError:
        print('Error')
        return False
    
def readData(filename, label):
    wb = px.load_workbook(filename, use_iterators = True)
    worksheets = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(worksheets[0])#'SK ID_129186 VoIP performance_n')
    cnt = 0
    ind = -1
    for row in ws.iter_rows():
        print('cnt', cnt)
        for i in range(len(row)):
            if str(row[i].value) == label:
                ind = i
                print('found')
                break
        cnt += 1
        if cnt > 0:
            break
        
        if ind < 0:
            print('No such field')
            sys.exit(0)
    
    nr = 0 
    values = []       
    for row in ws.iter_rows():
        
        if nr == 0:
            nr += 1
            continue
        nr += 1
        v = row[ind].value
        if v is not None:
            #print(nr, v)
            values.append((nr, float(v)))
    return values
        
def getLengths(x, y, x_l, y_l, x_r, y_r):
    a = math.sqrt(pow((x-x_l), 2) + pow((y-y_l), 2))
    b = math.sqrt(pow((x-x_r), 2) + pow((y-y_r), 2))
    c = math.sqrt(pow((x_r-x_l), 2) + pow((y_r-y_l), 2))
    
    return (a, b, c)
            
def getAvgCosine(x, y, points, s):   
    s_l = min(x, s) 
    s_r = min(len(points) -x, s)
    sample_left = rnd.sample(range(x), s_l)
    sample_right = rnd.sample(range(x, len(points)), s_r)
    sample = min(s_l, s_r)
    cosine = 0
    cnt = 0
    for ind in range(sample):
        (x_l, y_l) = points[sample_left[ind]]
        (x_r, y_r) = points[sample_right[ind]]
        (a, b, c) = getLengths(x, y, x_l, y_l, x_r, y_r)
        if a == 0 or b == 0:
            continue
        cos = (a*a + b*b - c*c)/(2*a*b)
        cnt += 1
        cosine += cos
    return cosine/cnt   
    
def getAvgCosineSW(x, y, points, s, rng):
    s_l = min(x, rng) 
    s_r = min(len(points) -x, rng)
    if s_l < 200 or s_r <200:
        return -1
    cosine = 0
    cnt = 0
    for _ in range(s):
        ind1 = rnd.randrange(x-s_l, x)
        ind2 = rnd.randrange(x, x+s_r)
        (x_l, y_l) = points[ind1]
        (x_r, y_r) = points[ind2]
        (a, b, c) = getLengths(x, y, x_l, y_l, x_r, y_r)
        if a == 0 or b == 0:
            continue
        cos = (a*a + b*b - c*c)/(2*a*b)
        cnt += 1
        cosine += cos
    return cosine/cnt   
        
        
def getOutliers(points, size, samples):
    heap = []
    cnt = 0
    for (x, y) in points:
        cnt += 1
        if cnt < 10 or cnt > len(points) - 10:
            continue
        if cnt % 10000 == 0:
            print(cnt)
        ang = getAvgCosine(x, y, points, samples)
        if len(heap) < size or ang > heap[0][0]:
            heapq.heappush(heap, (ang, (x, y)))
            if len(heap) > size:
                heapq.heappop(heap)   
   
    return heap

    
    
def readFromTextFile(filename):
    fo = open(filename, 'r')
    values = []
    nr = 1
    for val in fo:
        v = float(val)
        values.append((nr, v))
        nr += 1
    return values

def getOutliersSW(points, size, samples, rng):
    heap = []
    cnt = 0
    for (x, y) in points:
        if cnt % 10000 == 0:
            print(cnt)
        cnt += 1
        cos = getAvgCosineSW(x, y, points, samples, rng)
        if len(heap) < size or cos > heap[0][0]:
            heapq.heappush(heap, (cos, (x, y)))
            if len(heap) > size:
                heapq.heappop(heap)      
    return heap

def getWeek(values, outliers, mval, nr):
    i1 = 1440*nr#int(len(values)*((nr-1)/4.0))
    i2 = 1440*(nr+1)#int(len(values)*(nr/4.0))
    print(i1, i2)
    vals = values[i1:i2]
    outliers_ind = []
    outliers_val = []
    for (ind, val) in outliers:
        if ind in range(i1, i2):
            outliers_ind.append(ind-i1)
            outliers_val.append(val)
    print(len(outliers_ind))
    outliers_val_norm = [v/mval for v in outliers_val]
    values_norm = [v/mval for v in vals]
    return (values_norm, outliers_ind, outliers_val_norm)

def plotOutliers(values, outliers, days=3):
    _, axarr = plt.subplots(days, sharex = True, sharey = True)
    mval = max(values)
    offset = 0
    for d in range(0, days):
        print('day ' + str(d))
        vals, outliers_ind, outliers_val = getWeek(values, outliers, mval, d + offset)
        #axarr[w].set_yticks([0,0.5,1])
        axarr[d].plot(vals)
        axarr[d].plot(outliers_ind, outliers_val, u'ro')
        #axarr[cnt].plot(Xrec[normday[cnt],:])
        #axarr[cnt].set_ylim([minnx,maxnx])
        #axarr[cnt].set_yticklabels(['0','0.5','1'])
#      
        #plt.legend(['Real Latency','Low-rank approx.'],bbox_to_anchor=(1, -0.5))
        #plt.text(500,25.5,'Regular Days', fontsize=14, fontweight ='bold')
        plt.xlabel('Time', fontsize=14, fontweight ='bold')
        axarr[d].set_ylim([0, 1.2])
        axarr[d].set_xlim([1, len(vals)])
    font = {'weight' : 'bold','size'   : 12}
    plt.rc('font',**font)
    plt.show()
    
if __name__ == "__main__":
    filename = '/home/koki/Desktop/darepo/BT/SK ID_129186 VoIP performance_14nov-13dec2014.xlsx'
    txtfile =  '/home/koki/Desktop/darepo/BT/Latency1m.txt'
    points = readFromTextFile(txtfile)#readData(filename, 'dtime')
    print(len(points))
    top = 300
    samples = 100
    rng = 500
    outliers = getOutliersSW(points, top, samples, rng)
    for (ang, (x, y)) in outliers:
        print(ang, x, y)
    indices = [x[0] for x in points]
    values = [x[1] for x in points]
    
    outliers2 = [x[1] for x in outliers]
    mval = max(values)
    plotOutliers(values, outliers2)
    sys.exit()
    outliers_ind = [x[1][0] for x in outliers]
    outliers_val = [x[1][1] for x in outliers]
    values_norm = [v/mval for v in values]
    outliers_val_norm = [v/mval for v in outliers_val]
    plt.plot(indices, values_norm)
    plt.plot(outliers_ind, outliers_val_norm, u'ro')
    plt.xlabel('Time', fontsize=14, fontweight ='bold')
    plt.plot()
    plt.ylim([0,1.2])
    plt.show()
    #for k in outliers:
    #    print(k)
#     values = [x[1] for x in points]
#     plt.plot(values)
#     plt.show()